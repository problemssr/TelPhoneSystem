import sqlite3
from openpyxl import load_workbook
import os

conn = sqlite3.connect("../TelPhoneSystem/phone.sqlite3")
cur = conn.cursor()
# media_path=os.path.join("media")
# sheet = phone['Sheet1']
# print(phone.worksheets[0])
load_path = ""
phone=[]
for i in os.listdir("media"):
    if i.__contains__("上架手机"):
        load_path = os.path.join("media", i)
    phone = load_workbook(load_path)
    # sheet = phone['Sheet1']
sheet=phone.worksheets[0]
print(phone.worksheets[0])

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6):
    cur.execute(
        "insert into phone_info (phone_id,phone_name,phone_brand,phone_serial_number,phone_source,phone_IMEI) values (?,?,?,?,?,?)",
        (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))
    conn.commit()

print("Successful!!")

cur.close()
