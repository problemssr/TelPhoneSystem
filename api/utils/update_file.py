import datetime
import os
import sqlite3
from openpyxl import load_workbook


class Update_file(object):

    def upload_phonelist(self):
        conn = sqlite3.connect("../TelPhoneSystem/phone.sqlite3")
        cur = conn.cursor()
        load_path = ""
        for i in os.listdir("media"):
            if i.__contains__("序列号"):
                load_path = os.path.join("media", i)

        phone = load_workbook(load_path)

        sheet = phone.worksheets[0]
        # phone = load_workbook('media/序列号清单20220808.xlsx')
        # sheet = phone.worksheets[0]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6):
            cur.execute(
                "insert into phone_info (phone_id,phone_name,phone_brand,phone_serial_number,phone_source,phone_IMEI) values (?,?,?,?,?,?)",
                (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))
            conn.commit()

        # print("Successful1!!")

        cur.close()
        return 0

    def upload_phoneInfolist(self):
        conn = sqlite3.connect("../TelPhoneSystem/phone.sqlite3")
        cur = conn.cursor()

        load_path = ""
        for i in os.listdir("media"):
            if i.__contains__("手机资料"):
                load_path = os.path.join("media", i)

        phone = load_workbook(load_path)
        sheet = phone.worksheets[0]

        # 107列
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=108):
            cur.execute(
                "insert into detail_list ("
                "detail_id,detail_name,detail_storage,detail_storage_id,detail_brand,detail_price,"
                "detail_buyer,detail_comment,detail_comment_num,detail_exposure_date,detail_phone_type,"
                "detail_CVT,detail_Touchscreen_type,detail_screen_size,detail_screen_material,detail_screen_resolution,"
                "detail_screen_pixel,detail_narrow_bezel,detail_screen_of,detail_screen_technology,detail_network_type,"
                "detail_3G_network,detail_2G_spectrum,detail_3GTDSCDMA,detail_3GWCDMA,detail_3GCDMA,detail_4G,"
                "detail_WLAN_function,detail_nav,detail_connection_sharing,detail_NFC,detail_Bluetooth40,"
                "detail_OTG,detail_IR,detail_MHL,detail_Bluetooth,detail_uplink_rate,detail_downlink_rate,"
                "detail_HCE,detail_HCE_zone,detail_4G_network,detail_OS,detail_Core,detail_CPU_freq,"
                "detail_RAM,detail_ROM,detail_SD,detail_GB,detail_Cell,detail_MAH,detail_CPU_type,"
                "detail_GPU_type,detail_user_views,detail_talk_time,detail_stand_time,detail_camera,"
                "detail_camera_type,detail_back_pixel,detail_front_pixel,detail_CMOS,detail_shooting,"
                "detail_Photograph,detail_flash,detail_Aperture,detail_stylist,detail_Color,detail_Mobile_size,"
                "detail_Mobile_weight,detail_Body_Material,detail_Operation_type,detail_sensor_type,detail_SIM,"
                "detail_fuselage_interface,detail_fuselage_characteristics,detail_audio,detail_video,detail_picture,"
                "detail_common_function,detail_business_function,detail_service_features,detail_list1,"
                "detail_list2,detail_list3,detail_list4,detail_list5,detail_list6,detail_list7,detail_Novero,"
                "detail_quality_time,detail_quality_remark,detail_service_tel,detail_tel_remark,detail_field90,"
                "detail_ROOT,detail_ROOT_remark,detail_Adb,detail_device_number,detail_MAC_address,"
                "detail_remark,detail_power_specification,detail_power_manufacturer,detail_purchase_date,"
                "detail_Model,detail_alias,detail_IMEI,detail_charging,detail_wireless_charging) "
                "values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (row[0].value, row[1].value, row[2].value, int(row[3].value), row[4].value, row[5].value,
                 row[6].value, row[7].value, row[8].value, row[9].value, row[10].value, row[11].value,
                 row[12].value, row[13].value, row[14].value, row[15].value, row[16].value, row[17].value,
                 row[18].value, row[19].value, row[20].value, row[21].value, row[22].value, row[23].value,
                 row[24].value, row[25].value, row[26].value, row[27].value, row[28].value, row[29].value,
                 row[30].value, row[31].value, row[32].value, row[33].value, row[34].value, row[35].value,
                 row[36].value, row[37].value, row[38].value, row[39].value, row[40].value, row[41].value,
                 row[42].value, row[43].value, row[44].value, row[45].value, row[46].value, row[47].value,
                 row[48].value, row[49].value, row[50].value, row[51].value, row[52].value, row[53].value,
                 row[54].value, row[55].value, row[56].value, row[57].value, row[58].value, row[59].value,
                 row[60].value, row[61].value, row[62].value, row[63].value, row[64].value, row[65].value,
                 row[66].value, row[67].value, row[68].value, row[69].value, row[70].value, row[71].value,
                 row[72].value, row[73].value, row[74].value, row[75].value, row[76].value, row[77].value,
                 row[78].value, row[79].value, row[80].value, row[81].value, row[82].value, row[83].value,
                 row[84].value, row[85].value, row[86].value, row[87].value, row[88].value, row[89].value,
                 row[90].value, row[91].value, row[92].value, row[93].value, row[94].value, row[95].value,
                 row[96].value, row[97].value, row[98].value, row[99].value, row[100].value, row[101].value,
                 row[102].value, row[103].value, row[104].value, row[105].value, row[106].value))
            conn.commit()
        cur.close()
        return 0
