import datetime
import os
import sqlite3

import openpyxl
from openpyxl import load_workbook

from django import forms
from django.http import FileResponse
from django.shortcuts import render, redirect

from api import models
from api.utils.update_file import Update_file
from api.utils.pagenation import Pagenation


# Create your views here.
def phone_list(request):
    # 搜索----start
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"phone_brand__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.phone_info.objects.filter(**data_dict)
    count = models.phone_info.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "联迪移动支付实验室设备一览表",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }
    return render(request, 'phone_list.html', context)


def user_borrow(request):
    # 搜索----start
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"phone_brand__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.phone_info.objects.filter(**data_dict)
    count = models.phone_info.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "联迪移动支付实验室设备一览表",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }
    return render(request, 'user_borrow.html', context)


"""直接去首页"""


def to_index(request):
    return redirect('/phone/list/')


class DetailModelForm(forms.ModelForm):
    class Meta:
        model = models.detail_list
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.count = 1
        for name, field in self.fields.items():
            self.count += 1


def phone_detail(request, nid):
    """特定手机详情"""
    detail_data = models.detail_list.objects.filter(detail_id=nid).first()
    # print(detail_data.detail_purchase_date)
    form = DetailModelForm(instance=detail_data)
    context = {
        "title": "手机详细信息",
        "forms": form,
        "count": "具体参数：" + str(form.count) + " 项",
    }
    return render(request, "phone_detail.html", context)


def phone_instructions(request):
    """使用说明"""
    return render(request, "phone_instructions.html")


"""近一个月新增手机"""


def one_month_list(request):
    now = datetime.datetime.now()
    thiryDays = datetime.timedelta(days=31)  # 获取30天前的日期
    checkDate = (now - thiryDays).strftime('%Y-%m-%d %H:%M:%S')
    detail = models.detail_list.objects.filter(detail_purchase_date__gte=checkDate)
    data_list = []
    count = 0

    for d in detail:
        data_list.append(d.detail_id)
        count += 1

    data = models.phone_info.objects.filter(phone_id__in=data_list)
    print(data)
    context = {
        "title": "近一个月新增设备",
        "forms": data,
        "count": "近一个月新增设备：" + str(count) + " 台",
    }
    return render(request, "one_month_list.html", context)


"""筛选手机"""


def select_list(request):
    details = models.detail_list.objects.all()
    brand = set()
    phone_type = set()
    CVT = set()
    screen_size = set()
    screen_resolution = set()
    screen_pixel = set()
    NFC = set()
    Bluetooth40 = set()
    OTG = set()
    IR = set()
    MHL = set()
    Bluetooth = set()
    HCE = set()
    OS = set()
    CPU = set()
    SIM = set()
    for detail in details:
        brand.add(detail.detail_brand)
        if detail.detail_phone_type != None:
            phone_type.add(detail.detail_phone_type)
        if detail.detail_CVT != None:
            CVT.add(detail.detail_CVT)
        if detail.detail_screen_size != None:
            screen_size.add(detail.detail_screen_size)

        if detail.detail_screen_resolution != None:
            screen_resolution.add(detail.detail_screen_resolution)

        if detail.detail_screen_pixel != None:
            screen_pixel.add(detail.detail_screen_pixel)

        if detail.detail_NFC != None:
            NFC.add(detail.detail_NFC)

        if detail.detail_Bluetooth40 != None:
            Bluetooth40.add(detail.detail_Bluetooth40)

        if detail.detail_OTG != None:
            OTG.add(detail.detail_OTG)

        if detail.detail_IR != None:
            IR.add(detail.detail_IR)

        if detail.detail_MHL != None:
            MHL.add(detail.detail_MHL)

        if detail.detail_Bluetooth != None:
            Bluetooth.add(detail.detail_Bluetooth)

        if detail.detail_HCE_zone != None:
            HCE.add(detail.detail_HCE_zone)

        if detail.detail_OS != None:
            OS.add(detail.detail_OS)

        if detail.detail_CPU_type != None:
            CPU.add(detail.detail_CPU_type)

        if detail.detail_SIM != None:
            SIM.add(detail.detail_SIM)
    if request.method == "GET":
        context = {
            "title": "手机筛选",
            "brand": brand,
            "phone_type": phone_type,
            "CVT": CVT,
            "screen_size": screen_size,
            "screen_resolution": screen_resolution,
            "screen_pixel": screen_pixel,
            "NFC": NFC,
            "Bluetooth40": Bluetooth40,
            "OTG": OTG,
            "IR": IR,
            "MHL": MHL,
            "Bluetooth": Bluetooth,
            "HCE": HCE,
            "OS": OS,
            "CPU": CPU,
            "SIM": SIM,
        }
        return render(request, "select_list.html", context)
    else:
        res = request.POST
        data_dict = {
            'detail_brand__contains': res['brand'],
            'detail_phone_type__contains': res['phone_type'],
            'detail_CVT__contains': res['CVT'],
            'detail_screen_size__contains': res['screen_size'],
            'detail_screen_resolution__contains': res['screen_resolution'],
            'detail_screen_pixel__contains': res['screen_pixel'],
            'detail_NFC__contains': res['NFC'],
            'detail_Bluetooth40__contains': res['Bluetooth40'],
            'detail_OTG__contains': res['OTG'],
            'detail_IR__contains': res['IR'],
            'detail_Bluetooth__contains': res['Bluetooth'],
            'detail_HCE_zone__contains': res['HCE'],
            'detail_OS__contains': res['OS'],
            'detail_CPU_type__contains': res['CPU'],
            'detail_SIM__contains': res['SIM'],
            'detail_MHL__contains': res['MHL'],
        }
        data_list = []
        count = 0
        aid = models.detail_list.objects.filter(**data_dict)
        for d in aid:
            data_list.append(d.detail_id)
            count += 1
        data = models.phone_info.objects.filter(phone_id__in=data_list)

        context = {
            "title": "手机筛选结果",
            "count": "筛选结果：" + str(count) + "台",
            "forms": data
        }
        return render(request, "select_submit.html", context)


"""下载数据"""


def download(request, flag):
    for i in os.listdir("media"):
        if flag == 1:
            for j in os.listdir("media"):
                if j.__contains__("手机资料"):
                    f = open(os.path.join("media", j), "rb")
                    r = FileResponse(f, as_attachment=True, filename=j)
                    return r
        elif flag == 2:
            # 数据库写入Excel
            for i in os.listdir("media"):
                if i.__contains__("序列号"):
                    os.remove(os.path.join("media", i))

            info = models.phone_info.objects.all()
            # 新建一个工作簿
            wb = openpyxl.Workbook()
            # 新建一个工作表
            ws = wb.create_sheet('Sheet1', 0)
            # 表的首行
            ws.append(['ID', '设备类型(型号)', '品牌', '序列号', '来源', 'IMEI'])
            for dic in info:
                # 将数据库查询到的数据写到excel表格中
                ws.append(
                    [dic.phone_id, dic.phone_name,
                     dic.phone_brand, dic.phone_serial_number,
                     dic.phone_source, dic.phone_IMEI]
                )
            # 文件名
            name = datetime.datetime.now().strftime('%Y%m%d') + "序列号清单" + ".xlsx"
            # 这个是media文件夹下的students_excels文件夹路径，保存学生消息的excel表
            PATH = os.path.join('media', name)
            wb.save(PATH)
            f = open(os.path.join("media", name), "rb")
            r = FileResponse(f, as_attachment=True, filename=name)
            return r
        else:
            return render(request, "failed.html")


"""登录(使用Form)"""


class LoginForm(forms.Form):
    # Form自定义字段
    username = forms.CharField(
        label="用户名",
        widget=forms.TextInput,
        required=True
    )
    password = forms.CharField(
        label="密码",
        widget=forms.PasswordInput(render_value=True),
        required=True
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {
                "class": "form-control",
                "placeholder": field.label
            }


def login(request):
    if request.method == "GET":
        form = LoginForm()
        return render(request, "login.html", {"form": form})
    else:
        form = LoginForm(data=request.POST)
        if form.is_valid():
            # print(form.cleaned_data)
            data = models.admin.objects.filter(**form.cleaned_data).first()
            if not data:
                form.add_error("password", "用户名或密码错误")
                return render(request, "login.html", {"form": form})
            # 存session中
            request.session["info"] = {
                "id": data.id,
                "username": data.username
            }
            # 保存session7天
            request.session.set_expiry(60 * 60 * 24 * 7)
            return redirect('/admin/list/')
        return render(request, "login.html", {"form": form})


"""注销"""


def logout(request):
    """注销"""
    request.session.clear()

    return redirect('/login/')


"""管理员主界面"""


def admin_list(request):
    context = {
        "title": "手机借还登记"
    }
    return render(request, 'adminHTML/index.html', context)


"""借出登记"""


def admin_borrow(request):
    # 搜索----start
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"phone_brand__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.phone_info.objects.filter(**data_dict).all()

    count = models.phone_info.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "借出登记",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }
    return render(request, 'adminHTML/borrow.html', context)


def admin_borrow_detail(request, nid, user):
    data = models.phone_info.objects.filter(phone_id=nid).first()
    if user:
        title = "用户借用登记"
    else:
        title = "借出设备登记"
    context = {
        "title": title,
        "form_phone": data,
    }
    # print(data.user.people_name)
    if request.method == "GET":
        if user:
            return render(request, 'userHTML/borrow_detail.html', context)
        else:
            return render(request, 'adminHTML/borrow_detail.html', context)
    else:
        people_name = request.POST.get('borrow_name')
        restore_time = request.POST.get('restore_time')
        borrow_usage = request.POST.get('borrow_usage')
        phone = request.POST.get('phone')
        # 存入借用表
        models.borrow_info.objects.create(people_name=people_name,
                                          restore_time=restore_time,
                                          borrow_usage=borrow_usage,
                                          phone=phone)
        # 修改手机表
        models.phone_info.objects.filter(phone_id=phone).update(user=phone)

        # 插入历史记录表
        models.borrow_history.objects.create(history_name=data.phone_name,
                                             history_brand=data.phone_brand,
                                             history_phone_id=data.phone_id,
                                             history_source=data.phone_source,
                                             history_IMEI=data.phone_IMEI,
                                             history_username=people_name,
                                             history_borrow_time=datetime.datetime.now().strftime("%Y-%m-%d"),
                                             history_restore_time=restore_time,
                                             history_usage=borrow_usage)
        if user:
            return redirect('/phone/borrowInfo/')
        else:
            return redirect('/admin/borrow/')


"""归还登记"""


def admin_restore(request):  # 搜索----start
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"phone_brand__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.phone_info.objects.filter(**data_dict)
    count = models.phone_info.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "归还登记",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }

    return render(request, 'adminHTML/restore.html', context)


def admin_restore_detail(request, nid, user):
    data = models.phone_info.objects.filter(phone_id=nid).first()
    if user:
        title = "用户归还设备"
    else:
        title = "确定归还设备"
    context = {
        "title": title,
        "form_phone": data,
    }
    if user:
        return render(request, 'userHTML/restore_detail.html', context)
    else:
        return render(request, 'adminHTML/restore_detail.html', context)


def delete_restore_detail(request, nid, user):
    """删除用户表记录"""
    models.borrow_info.objects.filter(phone=nid).delete()
    if user:
        return redirect('/phone/borrowInfo/')
    else:
        return redirect('/admin/restore/')


class PhoneAddModelForm(forms.ModelForm):
    class Meta:
        model = models.phone_info
        exclude = ["user"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            field.widget.attrs = {
                "class": "form-control", "placeholder": field.label
            }


"""新增设备"""


def admin_add(request):
    if request.method == "GET":
        getforms = PhoneAddModelForm()
        context = {
            "title": "新增设备",
            "form_phone": getforms,
        }
        return render(request, 'adminHTML/add_detail.html', context)
    else:
        postforms = PhoneAddModelForm(data=request.POST)
        if postforms.is_valid():
            postforms.save()
            return redirect('/admin/list/')
        context = {
            "title": "新增设备",
            "form_phone": postforms,
        }
        return render(request, 'adminHTML/add_detail.html', context)


"""编辑设备"""


def admin_edit(request):
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"phone_brand__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.phone_info.objects.filter(**data_dict)
    count = models.phone_info.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "编辑设备信息",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }

    return render(request, 'adminHTML/edit.html', context)


class PhoneModelForm(forms.ModelForm):
    class Meta:
        model = models.phone_info
        fields = ['phone_id', 'phone_brand', 'phone_name', 'phone_source', 'phone_IMEI']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            field.widget.attrs = {
                "class": "form-control", "placeholder": field.label
            }


def admin_edit_detail(request, nid):
    data = models.phone_info.objects.filter(phone_id=nid).first()
    if request.method == "GET":
        forms = PhoneModelForm(instance=data)
        context = {
            "title": "更新设备",
            "form_phone": forms,
        }
        return render(request, 'adminHTML/edit_detail.html', context)
    else:
        forms = PhoneModelForm(data=request.POST, instance=data)
        if forms.is_valid():
            forms.save()
            models.detail_list.objects.filter(detail_id=nid).update(
                detail_brand=forms.data.get('phone_brand'),
                detail_name=forms.data.get('phone_name'),
                detail_IMEI=forms.data.get('phone_IMEI')
            )
            return redirect('/admin/edit/')
        context = {
            "title": "更新设备",
            "form_phone": forms,
        }
        return render(request, 'adminHTML/edit_detail.html', context)


"""借出设备历史记录"""


def admin_histroy(request):
    # 搜索----start
    data_dict = {}
    search_data = request.GET.get('search', "")
    if search_data:
        data_dict = {"history_name__contains": search_data}
    # 搜索----end

    # 所有数据
    data = models.borrow_history.objects.filter(**data_dict).all()

    count = models.borrow_history.objects.count()
    page_size = request.GET.get("page_size", 10)
    # 分页
    page_project = Pagenation(request, data, page_size)
    context = {
        "title": "借用手机历史记录",
        "count": "设备总数：" + str(count) + "台",
        "form": page_project.page_queryset,  # 分完页的数据
        "page_string": page_project.html()  # 生成页码
    }
    return render(request, 'adminHTML/histroy.html', context)


"""开发人员管理"""


def admin_manage(request):
    if request.method == "GET":
        return render(request, 'adminHTML/manage.html')
    else:
        """上传excel文件"""
        phone_obj = request.FILES.get('phone_data')
        # list_obj = request.FILES.get('list_data')

        # 清空media下面的文件
        for i in os.listdir("media"):
            # if i.__contains__("序列号"):
            #     os.remove(os.path.join("media", i))
            if i.__contains__("手机资料"):
                os.remove(os.path.join("media", i))
            else:
                pass

        # 保存到文件目录下
        infoname = datetime.datetime.now().strftime('%Y%m%d') + "手机资料" + ".xlsx"
        # listname = datetime.datetime.now().strftime('%Y%m%d') + "序列号清单" + ".xlsx"

        # 手机详细信息
        f = open(os.path.join('media', infoname), mode="wb")
        for chunk in phone_obj.chunks():
            f.write(chunk)
        f = open(os.path.join('excelRecord', infoname), mode="wb")
        for chunk in phone_obj.chunks():
            f.write(chunk)

        # 手机序列号
        # f = open(os.path.join('media', listname), mode="wb")
        # for chunk in list_obj.chunks():
        #     f.write(chunk)
        # f = open(os.path.join('excelRecord', listname), mode="wb")
        # for chunk in phone_obj.chunks():
        #     f.write(chunk)
        f.close()
        # 更新数据库
        upload = Update_file()

        # upload.upload_phonelist()
        upload.upload_phoneInfolist()

        models.borrow_info.objects.all().delete()

        return redirect('/admin/manage/')


def admin_delete(request):
    """删除数据"""
    # models.phone_info.objects.all().delete()
    models.detail_list.objects.all().delete()
    return redirect('/admin/manage/')
